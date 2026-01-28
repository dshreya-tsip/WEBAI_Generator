import os
import re
import docx
import openpyxl
import requests

# -------------------------------
# Step 1: Extract text from SRS.docx
# -------------------------------
def extract_srs_text(doc_path: str) -> str:
    doc = docx.Document(doc_path)
    return "\n".join([para.text for para in doc.paragraphs if para.text.strip()])


# -------------------------------
# Step 2: Build prompt for Claude
# -------------------------------
def build_prompt(srs_text: str) -> str:
    return (
        "Read the uploaded Software Requirements Specification (SRS.docx).\n"
        "You MUST output exactly two parts in this order:\n"
        "1) A single line in the exact format:\n"
        "   Component: <detected overall component/module/system name from the SRS>\n"
        "   (Put only this line first. No code fences, no extra text before it.)\n"
        "2) A blank line, followed immediately by a single markdown table of test cases.\n\n"

        "⚠️ IMPORTANT: Generate the **maximum possible coverage of test cases** from the SRS.\n"
        "- Include **all functional test cases** (for every requirement, feature, rule, and exception).\n"
        "- Include **all non-functional test cases**:\n"
        "  • Performance\n"
        "  • Usability\n"
        "  • Security\n"
        "  • Reliability\n"
        "  • Compatibility\n"
        "  • Accessibility\n"
        "  • Compliance\n"
        "  • Installation\n"
        "  • Recovery\n"
        "- Include **negative test cases** (invalid inputs, boundary conditions, failure handling).\n"
        "- Include **edge cases, stress cases, and corner cases**.\n"
        "- Include **ad-hoc / exploratory test cases** (unplanned scenarios, random inputs, unusual user flows).\n"
        "- Include **data validation test cases** (e.g., input formats, required fields, constraints).\n"
        "- Include **integration test cases** (interactions between modules/components).\n"
        "- Include **regression test cases** (to verify previously working features still work).\n"
        "- Include **accessibility test cases** (screen reader support, keyboard navigation, contrast ratios).\n"
        "- Do not skip any scenario implied in the SRS, even if not explicitly written.\n\n"

        "🚨 MANDATORY REQUIREMENT:\n"
        "- You MUST include a **separate set of test cases dedicated to IPv4 and IPv6**.\n"
        "- Cover at least the following:\n"
        "  • IPv4 only scenarios\n"
        "  • IPv6 only scenarios\n"
        "  • Dual-stack (IPv4 + IPv6) scenarios\n"
        "  • Fallback/switchover between IPv4 and IPv6\n"
        "  • Invalid/edge IP addresses (e.g., 0.0.0.0, 255.255.255.255, ::, ::1, malformed addresses)\n"
        "  • Performance comparison IPv4 vs IPv6\n"
        "  • Security checks for both IPv4 and IPv6\n\n"

        "✅ You MUST generate **at least 200 test cases** if the SRS is moderately detailed.\n"
        "If the SRS is short, extrapolate plausible scenarios based on typical systems.\n\n"

        "Number test cases sequentially across all categories with IDs like `TC001`, `TC002`, etc.\n"
        "All test cases must be in ONE continuous markdown table with no breaks or section headers.\n\n"

        "Return the markdown table with columns exactly named:\n"
        "`Test Case ID` | `Preconditions` | `Test Condition` | `Steps with description` | "
        "`Expected Result` | `Actual Result` | `Remarks`\n\n"

        "Notes for the header block in the Excel sheet (handled by my program):\n"
        "- The line you output as 'Component: <name>' will be written into the header's Component field.\n"
        "- `Build`, `Date`, and `Target` will remain blank.\n\n"

        "SRS Content:\n" + srs_text
    )





# -------------------------------
# Step 3: Send prompt to Claude API
# -------------------------------
def get_testcases_from_claude(srs_text: str) -> str:
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        raise ValueError("Missing Anthropic API key. Set ANTHROPIC_API_KEY environment variable.")

    prompt = build_prompt(srs_text)

    headers = {
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }

    payload = {
        "model": "claude-3-7-sonnet-20250219",
        "max_tokens": 8000,
        "temperature": 0.3,
        "messages": [
            {"role": "user", "content": prompt}
        ],
    }

    resp = requests.post("https://api.anthropic.com/v1/messages", json=payload, headers=headers, timeout=120)
    resp.raise_for_status()
    result = resp.json()

    md_full_text = "\n".join(
        block["text"] for block in result.get("content", []) if block.get("type") == "text"
    )

    print("\n--- Claude Raw Output (first 1200 chars) ---\n")
    print(md_full_text[:1200])
    print("\n--------------------------------------------\n")

    return md_full_text


# -------------------------------
# Step 4: Extract "Component: <name>"
# -------------------------------
def extract_component(md_full_text: str) -> str:
    m = re.search(r"(?im)^\\s*Component\\s*:\\s*(.+?)\\s*$", md_full_text)
    if m:
        return m.group(1).strip()
    return "Unknown"


# -------------------------------
# Step 5: Parse the markdown table
# -------------------------------
def parse_markdown_table(md_full_text: str):
    def clean_cell(value: str) -> str:
        if not value:
            return ""
        return value.replace("<br>", "\n").replace("\\n", "\n").strip()

    lines = md_full_text.splitlines()
    start_idx = None
    for i, line in enumerate(lines):
        if "|" in line and "Test Case ID" in line:
            start_idx = i
            break

    if start_idx is None:
        raise ValueError("Markdown table header not found in model output.")

    table_lines = []
    for line in lines[start_idx:]:
        if "|" in line:
            table_lines.append(line)
        else:
            if table_lines:
                break

    if len(table_lines) < 3:
        raise ValueError("Markdown table appears incomplete.")

    header_cells = [h.strip() for h in table_lines[0].split("|")[1:-1]]
    test_cases = []

    for row_line in table_lines[2:]:
        parts = [clean_cell(p) for p in row_line.split("|")[1:-1]]
        if len(parts) == len(header_cells):
            test_cases.append(dict(zip(header_cells, parts)))

    if not test_cases:
        raise ValueError("No test case rows parsed from the markdown table.")

    return test_cases


# -------------------------------
# Step 6: Write into Excel template
# -------------------------------
def fill_excel_template(test_cases, template_path: str, output_path: str, component_name: str):
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Testcases"]

    def set_header_field(label: str, value: str, search_rows: int = 10, search_cols: int = 12) -> bool:
        label_low = label.lower().rstrip(":")
        for r in range(1, search_rows + 1):
            for c in range(1, search_cols + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, str):
                    text = cell.value.strip()
                    if text.lower().startswith(label_low + ":"):
                        prefix = text.split(":", 1)[0]
                        cell.value = f"{prefix}: {value}".strip()
                        return True
        return False

    # Update only the Component field
    if not set_header_field("Component", component_name):
        ws["E2"] = f"Component: {component_name}"

    start_row = 6
    for i, tc in enumerate(test_cases):
        row = start_row + i
        ws.cell(row=row, column=2, value=tc.get("Test Case ID"))
        ws.cell(row=row, column=3, value=tc.get("Preconditions"))
        ws.cell(row=row, column=4, value=tc.get("Test Condition"))
        ws.cell(row=row, column=5, value=tc.get("Steps with description"))
        ws.cell(row=row, column=6, value=tc.get("Expected Result"))
        ws.cell(row=row, column=7, value=tc.get("Actual Result"))
        ws.cell(row=row, column=8, value=tc.get("Remarks"))

    wb.save(output_path)


# -------------------------------
# Main
# -------------------------------
if __name__ == "__main__":
    srs_path = "SRS.docx"
    template_path = "TestCases_Template.xlsx"
    output_path = "Generated_TestCases.xlsx"

    srs_text = extract_srs_text(srs_path)
    md_full = get_testcases_from_claude(srs_text)
    component = extract_component(md_full)
    print(f"✅ Detected Component: {component}")
    test_cases = parse_markdown_table(md_full)
    fill_excel_template(test_cases, template_path, output_path, component)
    print(f"✅ Test cases generated successfully: {output_path}")
